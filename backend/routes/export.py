"""
Export Route — /api/export/*
Supports: PDF (ReportLab), Excel (openpyxl), WhatsApp share link
"""
import os
import uuid
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file

logger = logging.getLogger(__name__)
export_bp = Blueprint("export", __name__)


# ── PDF Export ────────────────────────────────────────────────

@export_bp.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    data = request.get_json() or {}
    messages = data.get("messages", [])
    title = data.get("title", "ArthaAI — Financial Consultation")
    profile = data.get("profile", {})

    if not messages:
        return jsonify({"error": "No messages to export"}), 400

    try:
        from config import Config
        out_path = os.path.join(Config.EXPORT_DIR, f"arthaai_{uuid.uuid4().hex[:8]}.pdf")
        _generate_pdf(messages, title, profile, out_path)
        return send_file(out_path, mimetype="application/pdf",
                         as_attachment=True, download_name="ArthaAI_Report.pdf")
    except Exception as e:
        logger.error(f"PDF export error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _generate_pdf(messages: list, title: str, profile: dict, out_path: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    saffron = colors.HexColor("#FF6B00")
    navy = colors.HexColor("#0D1B3E")
    story = []

    # Title
    story.append(Paragraph(
        f'<font color="#FF6B00"><b>ArthaAI</b></font> — Financial Consultation Report',
        ParagraphStyle("Title", fontSize=18, spaceAfter=6, alignment=TA_CENTER,
                       textColor=navy)
    ))
    story.append(Paragraph(
        f'Generated on {datetime.now().strftime("%d %B %Y, %I:%M %p IST")}',
        ParagraphStyle("Sub", fontSize=10, spaceAfter=4, alignment=TA_CENTER,
                       textColor=colors.grey)
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=saffron, spaceAfter=12))

    # User profile
    if profile:
        profile_data = []
        if profile.get("name"):
            profile_data.append(["Name", profile["name"]])
        if profile.get("city"):
            profile_data.append(["City", profile["city"]])
        if profile.get("monthly_income"):
            profile_data.append(["Monthly Income", f"₹{profile['monthly_income']:,}"])
        if profile.get("risk_appetite"):
            profile_data.append(["Risk Appetite", profile["risk_appetite"]])
        if profile_data:
            story.append(Paragraph("User Profile", styles["Heading2"]))
            t = Table(profile_data, colWidths=[4*cm, 12*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FFF0E6")),
                ("TEXTCOLOR", (0, 0), (0, -1), saffron),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ROWBACKGROUNDS", (1, 0), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#EEEEEE")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t)
            story.append(Spacer(1, 12))

    # Conversation
    story.append(Paragraph("Consultation Transcript", styles["Heading2"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#EEEEEE"), spaceAfter=8))

    user_style = ParagraphStyle("User", fontSize=10, backColor=colors.HexColor("#0D1B3E"),
                                 textColor=colors.white, leftIndent=50, rightIndent=0,
                                 spaceBefore=8, spaceAfter=4, borderPadding=8,
                                 borderRadius=4)
    ai_style = ParagraphStyle("AI", fontSize=10, backColor=colors.HexColor("#F5F5F5"),
                               textColor=colors.HexColor("#222222"), leftIndent=0,
                               rightIndent=50, spaceBefore=4, spaceAfter=8,
                               borderPadding=8)

    for msg in messages:
        role = msg.get("role", "user")
        content = msg.get("content", "").replace("\n", "<br/>")
        if role == "user":
            story.append(Paragraph(f"<b>You:</b> {content}", user_style))
        else:
            story.append(Paragraph(f"<b>ArthaAI:</b> {content}", ai_style))

    # Footer
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#EEEEEE")))
    story.append(Paragraph(
        "⚠️ Disclaimer: This report is for informational purposes only. "
        "Investments are subject to market risks. Please consult a SEBI-registered advisor or CA before investing.",
        ParagraphStyle("Footer", fontSize=8, textColor=colors.grey, spaceBefore=8)
    ))

    doc.build(story)


# ── Excel Export ──────────────────────────────────────────────

@export_bp.route("/api/export/excel", methods=["POST"])
def export_excel():
    data = request.get_json() or {}
    messages = data.get("messages", [])

    if not messages:
        return jsonify({"error": "No messages to export"}), 400

    try:
        from config import Config
        out_path = os.path.join(Config.EXPORT_DIR, f"arthaai_{uuid.uuid4().hex[:8]}.xlsx")
        _generate_excel(messages, data.get("profile", {}), out_path)
        return send_file(out_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="ArthaAI_Consultation.xlsx")
    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _generate_excel(messages: list, profile: dict, out_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ─ Sheet 1: Conversation ─
    ws = wb.active
    ws.title = "Consultation"
    saffron_fill = PatternFill("solid", fgColor="FF6B00")
    navy_fill = PatternFill("solid", fgColor="0D1B3E")
    light_fill = PatternFill("solid", fgColor="F5F5F5")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 22

    # Header
    ws.merge_cells("A1:C1")
    ws["A1"] = "ArthaAI — Financial Consultation"
    ws["A1"].font = Font(bold=True, size=14, color="FF6B00")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p IST')}"
    ws["A2"].font = Font(size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    ws.append([])
    headers = ["Role", "Message", "Timestamp"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = navy_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Messages
    now = datetime.now()
    for i, msg in enumerate(messages):
        role = msg.get("role", "user").capitalize()
        content = msg.get("content", "")
        row = [role, content, now.strftime("%d/%m/%Y %H:%M")]
        ws.append(row)
        row_idx = ws.max_row
        fill = light_fill if role == "User" else PatternFill("solid", fgColor="FFF0E6")
        for col in range(1, 4):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
        ws.row_dimensions[row_idx].height = min(max(len(content) // 4, 20), 120)

    # ─ Sheet 2: Profile ─
    if profile:
        ws2 = wb.create_sheet("User Profile")
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 40
        ws2.append(["ArthaAI — User Financial Profile"])
        ws2["A1"].font = Font(bold=True, size=13, color="FF6B00")
        ws2.append([])
        for key, val in profile.items():
            ws2.append([key.replace("_", " ").title(), str(val)])
            ws2.cell(ws2.max_row, 1).font = Font(bold=True)

    wb.save(out_path)


# ── WhatsApp Share ────────────────────────────────────────────

@export_bp.route("/api/export/whatsapp", methods=["POST"])
def whatsapp_share():
    """Generate a WhatsApp share link with a summary of the conversation."""
    data = request.get_json() or {}
    messages = data.get("messages", [])
    phone = data.get("phone", "")   # Optional: pre-fill recipient number

    if not messages:
        return jsonify({"error": "No messages to share"}), 400

    # Build a short summary text
    lines = ["*ArthaAI Financial Advice Summary* 🇮🇳\n"]
    for msg in messages[-6:]:   # Last 3 exchanges
        role = "You" if msg.get("role") == "user" else "ArthaAI"
        content = msg.get("content", "")[:300]
        if len(msg.get("content", "")) > 300:
            content += "..."
        lines.append(f"*{role}:* {content}\n")
    lines.append("_Generated by ArthaAI — Your Indian Finance Advisor_")

    text = "\n".join(lines)
    import urllib.parse
    encoded = urllib.parse.quote(text)
    base = f"https://wa.me/{phone}" if phone else "https://wa.me"
    url = f"{base}?text={encoded}"

    return jsonify({"whatsapp_url": url, "text_preview": text[:200]})
